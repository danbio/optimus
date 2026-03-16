import re
import unicodedata
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView

from .forms import ImportarPlanilhaForm, ProdutoForm
from .models import Produto

# ── CRUD ──────────────────────────────────────────────────────────────────────


class ProdutoListView(LoginRequiredMixin, ListView):
    model = Produto
    template_name = "estoque/produto_list.html"
    context_object_name = "produtos"
    paginate_by = 30

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.GET.get("q", "").strip()
        bu = self.request.GET.get("bu", "")
        segmento = self.request.GET.get("segmento", "")
        ativo = self.request.GET.get("ativo", "1")

        if q:
            filtro = Q(descricao__icontains=q) | Q(ean__icontains=q) | Q(familia__icontains=q)
            try:
                filtro |= Q(codigo=int(q))
            except (ValueError, TypeError):
                pass
            qs = qs.filter(filtro)
        if bu:
            qs = qs.filter(bu=bu)
        if segmento:
            qs = qs.filter(segmento=segmento)
        qs = qs.filter(ativo=(ativo != "0"))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        ctx["bu_filtro"] = self.request.GET.get("bu", "")
        ctx["segmento_filtro"] = self.request.GET.get("segmento", "")
        ctx["ativo_filtro"] = self.request.GET.get("ativo", "1")
        ctx["bus"] = Produto.objects.values_list("bu", flat=True).exclude(bu="").distinct().order_by("bu")
        ctx["segmentos"] = Produto.objects.values_list("segmento", flat=True).exclude(segmento="").distinct().order_by("segmento")
        ctx["total_ativos"] = Produto.objects.filter(ativo=True).count()
        ctx["total_inativos"] = Produto.objects.filter(ativo=False).count()
        return ctx


class ProdutoDetailView(LoginRequiredMixin, DetailView):
    model = Produto
    template_name = "estoque/produto_detail.html"
    context_object_name = "produto"


class ProdutoCreateView(LoginRequiredMixin, CreateView):
    model = Produto
    form_class = ProdutoForm
    template_name = "estoque/produto_form.html"

    def get_success_url(self):
        return reverse_lazy("estoque:detalhe", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, f"Produto «{form.instance.descricao}» cadastrado com sucesso.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo"] = "Novo Produto"
        ctx["acao"] = "Cadastrar"
        return ctx


class ProdutoUpdateView(LoginRequiredMixin, UpdateView):
    model = Produto
    form_class = ProdutoForm
    template_name = "estoque/produto_form.html"

    def get_success_url(self):
        return reverse_lazy("estoque:detalhe", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, f"Produto «{form.instance.descricao}» atualizado com sucesso.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo"] = f"Editar — {self.object.descricao}"
        ctx["acao"] = "Salvar alterações"
        return ctx


class ProdutoDeleteView(LoginRequiredMixin, DeleteView):
    model = Produto
    template_name = "estoque/produto_confirm_delete.html"
    success_url = reverse_lazy("estoque:lista")

    def form_valid(self, form):
        messages.success(self.request, f"Produto «{self.object.descricao}» removido com sucesso.")
        return super().form_valid(form)


# ── IMPORTAÇÃO DE PLANILHA ─────────────────────────────────────────────────────

_MAPA_COLUNAS = {
    # BU: "Unidade" é o nome real na tabela Intelbras
    "bu": ["bu", "business unit", "unidade", "unidade de negocio", "unidade negocio"],
    "segmento": ["segmento", "seg"],
    "familia": ["familia", "family", "fam"],
    # "Código Produto" é o nome real na tabela Intelbras
    "codigo": ["codigo", "codigo produto", "cod", "code", "cod.", "item", "ref", "referencia", "part number", "pn", "sku", "codigo do produto", "cod produto"],
    # "Descrição do Produto" é o nome real na tabela Intelbras
    "descricao": ["descricao", "descricao do produto", "description", "desc", "desc.", "produto", "nome", "nome do produto", "denominacao"],
    "ipi": ["ipi", "ipi (%)", "ipi(%)", "aliq ipi", "aliquota ipi"],
    "icms": ["icms", "icms (%)", "icms(%)", "aliq icms", "aliquota icms"],
    "psd": ["psd", "preco custo", "preco de custo", "custo", "preco fabrica", "p. fabrica"],
    "pscf": ["pscf", "preco venda", "preco de venda", "venda", "preco sugerido", "p. sugerido", "pvp", "pv"],
    # "Preço Referência" é o nome real na tabela Intelbras
    "preco_referencia": ["preco referencia", "preco referencia", "preco_referencia", "referencia", "controle de venda", "controle venda", "tipo preco"],
    # "Qtd. Múltipla" é o nome real na tabela Intelbras
    "qtd_multipla": ["qtd. multipla", "qtd multipla", "quantidade multipla", "qtd_multipla", "qtd.multipla", "multiplo", "multiplo venda"],
    "ncm": ["ncm", "ncm/sh", "cod ncm"],
    "ean": ["ean", "ean/gtin", "gtin", "codigo de barras", "barcode", "cod barras", "ean 13"],
    "observacoes": ["observacoes", "observacoes", "obs", "obs.", "observacao", "nota"],
}


def _norm(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s)


def _mapear_colunas(header_row):
    mapa = {}
    # 1ª passagem: correspondência exata com o nome do campo (ex: "psd" → campo psd)
    # Garante que colunas com nome exato prevalecem sobre aliases genéricos
    # que podem surgir antes na mesma planilha (ex: "Custo" aparece antes de "PSD"
    # na tabela Intelbras, mas "Custo" é fórmula enquanto "PSD" tem valor real).
    for idx, cell in enumerate(header_row):
        n = _norm(cell)
        if n in _MAPA_COLUNAS and n not in mapa:
            mapa[n] = idx
    # 2ª passagem: aliases para campos ainda não mapeados
    for idx, cell in enumerate(header_row):
        n = _norm(cell)
        for campo, aliases in _MAPA_COLUNAS.items():
            if n in aliases and campo not in mapa:
                mapa[campo] = idx
                break
    return mapa


def _dec(val):
    if val is None:
        return Decimal("0")
    try:
        return Decimal(str(val)).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0")


def _pct(val):
    """Converte percentual detectando formato decimal: 0.065 → 6.50, 12.5 → 12.50."""
    if val is None:
        return Decimal("0")
    try:
        v = Decimal(str(val))
        # Intelbras armazena como decimal (0.065 = 6.5%) — detecta e converte
        if Decimal("0") < v < Decimal("1"):
            v = v * 100
        return v.quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0")


def _ler_planilha(arquivo, nome):
    if nome.endswith(".xlsb"):
        try:
            from pyxlsb import open_workbook
        except ImportError:
            raise ImportError("Biblioteca pyxlsb não instalada. Execute: pip install pyxlsb")
        with open_workbook(arquivo) as wb:
            sheet_names = list(wb.sheets)
            # Tenta cada aba até encontrar uma com cabeçalho válido
            for sheet_name in sheet_names:
                with wb.get_sheet(sheet_name) as sheet:
                    rows = [[cell.v for cell in row] for row in sheet.rows()]
                idx, _ = _encontrar_cabecalho(rows)
                if idx is not None:
                    return rows
        raise ValueError(f'Nenhuma aba contém as colunas "Codigo" e "Descricao". Abas encontradas: {", ".join(sheet_names[:8])}.')
    else:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("Biblioteca openpyxl não instalada. Execute: pip install openpyxl")
        wb = load_workbook(arquivo, data_only=True, read_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            raise ValueError("Planilha sem abas ativas.")
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        wb.close()
        return rows


def _encontrar_cabecalho(rows):
    """Varre até a 50ª linha buscando a que contenha 'codigo' e 'descricao'.
    Aceita também linhas com apenas um dos dois (prioriza quem tem os dois)."""
    melhor = (None, None)
    for i, row in enumerate(rows[:50]):
        mapa = _mapear_colunas(row)
        if "codigo" in mapa and "descricao" in mapa:
            return i, mapa
        if melhor[0] is None and ("codigo" in mapa or "descricao" in mapa):
            melhor = (i, mapa)
    return melhor


def _processar_rows(rows):
    idx_header, mapa = _encontrar_cabecalho(rows)
    if idx_header is None:
        raise ValueError('Cabeçalho não encontrado. Verifique se a planilha contém as colunas "Codigo" e "Descricao".')

    criados = atualizados = erros = 0
    codigos_vistos: set[int] = set()

    for row in rows[idx_header + 1 :]:
        if not any(v is not None and str(v).strip() for v in row):
            continue

        def get(campo: str, _row: list = row, _mapa: dict = mapa):  # type: ignore[assignment]
            idx = _mapa.get(campo)
            if idx is None or idx >= len(_row):
                return None
            return _row[idx]

        try:
            codigo = int(float(str(get("codigo"))))
        except (TypeError, ValueError):
            erros += 1
            continue

        # Pula linhas duplicadas (mesmo produto repetido por UF/estado)
        if codigo in codigos_vistos:
            continue
        codigos_vistos.add(codigo)

        ean = str(get("ean") or "").strip()
        if ean in ("-", "N/A", "n/a", "—"):
            ean = ""

        ref = str(get("preco_referencia") or "").strip()
        # Normaliza valores conhecidos
        if ref.upper() == "PSD":
            ref = "PSD"
        elif ref.upper() == "PP":
            ref = "PP"

        try:
            _, created = Produto.objects.update_or_create(
                codigo=codigo,
                defaults={
                    "bu": str(get("bu") or "").strip(),
                    "segmento": str(get("segmento") or "").strip(),
                    "familia": str(get("familia") or "").strip(),
                    "descricao": str(get("descricao") or "").strip(),
                    "ipi": _pct(get("ipi")),
                    "icms": _pct(get("icms")),
                    "psd": _dec(get("psd")),
                    "pscf": _dec(get("pscf")),
                    "preco_referencia": ref,
                    "qtd_multipla": _dec(get("qtd_multipla")) or Decimal("1"),
                    "ncm": str(get("ncm") or "").strip(),
                    "ean": ean,
                    "observacoes": str(get("observacoes") or "").strip(),
                },
            )
            if created:
                criados += 1
            else:
                atualizados += 1
        except Exception:
            erros += 1

    return criados, atualizados, erros


@login_required
def importar_tabela(request):
    if request.method == "POST":
        form = ImportarPlanilhaForm(request.POST, request.FILES)
        if form.is_valid():
            arquivo = request.FILES["arquivo"]
            try:
                rows = _ler_planilha(arquivo, arquivo.name.lower())
                criados, atualizados, erros = _processar_rows(rows)
                partes = []
                if criados:
                    partes.append(f"{criados} criado(s)")
                if atualizados:
                    partes.append(f"{atualizados} atualizado(s)")
                if erros:
                    partes.append(f"{erros} linha(s) com erro")
                texto = " • ".join(partes) if partes else "Nenhum produto processado."
                messages.success(request, f"Importação concluída: {texto}.")
                return redirect("estoque:lista")
            except Exception as e:
                messages.error(request, f"Erro ao processar planilha: {e}")
    else:
        form = ImportarPlanilhaForm()

    colunas_info = [
        ("BU", "Unidade de negócio (ex: BU SEGURANCA)"),
        ("Segmento", "Segmento do produto"),
        ("Familia", "Família do produto"),
        ("Codigo", "Código único do produto *"),
        ("Descricao", "Descrição completa *"),
        ("IPI", "Alíquota IPI (%)"),
        ("ICMS", "Alíquota ICMS (%)"),
        ("PSD", "Preço de Custo"),
        ("PSCF", "Preço de Venda"),
        ("Preco_referencia", "Controle de venda (PSD / PP / S/Controle)"),
        ("Qtd_multipla", "Quantidade mínima por pedido"),
        ("NCM", "Código NCM fiscal"),
        ("EAN", "Código de barras EAN/GTIN"),
        ("Observacoes", "Observações (opcional)"),
    ]
    return render(request, "estoque/produto_importar.html", {"form": form, "colunas_info": colunas_info})
